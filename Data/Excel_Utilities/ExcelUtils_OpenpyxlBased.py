import os
import openpyxl
from Logic.EmployeeData import EmployeeData


def loadEmployeesFromExcel(employeesList, employeesFilePath, employeesFileName):

    workbook = openpyxl.load_workbook(os.path.join(employeesFilePath, employeesFileName))
    worksheet = workbook.get_sheet_by_name('EmployeesData')

    columnsName = readRowValues(worksheet, 1, 1,None)  # read the header.  The row where we stock the name of the column

    # tronsform every row of the worksheet into a  dict
    for row_index in range(1,worksheet.max_row ):
        columnValues = readRowValues(worksheet, row_index+1, 1, None)  # read the other rows
        newEmployeeData = dict(zip(columnsName, columnValues))

        employeesList.newEmployee(EmployeeData.createFromDict(newEmployeeData))

def insertRowInExcel(employeeData, employeesFilePath, employeesFileName):  # employeeData is a dictionary

    workbook = openpyxl.load_workbook(os.path.join(employeesFilePath, employeesFileName))
    worksheet = workbook.get_sheet_by_name('EmployeesData')

    jobIdToFind_RowIndex = _getRowIndexByEmployeeName(worksheet, employeeData['jobId'])  # 0 is the column index of the nombre

    #worksheet.max_row because we will insert at the end of the sheet
    _insertRowAtSheet(worksheet, worksheet.max_row + 1, employeeData) #because excel is 1 based and row is occupied by the headers
    workbook.save(os.path.join(employeesFilePath, employeesFileName))

def deleteRowInExcel(employeeData, employeesFilePath, employeesFileName):  # employeeData is a dictionary

    workbook = openpyxl.load_workbook(os.path.join(employeesFilePath, employeesFileName))
    worksheet = workbook.get_sheet_by_name('EmployeesData')
    jobIdToFind_RowIndex = _getRowIndexByEmployeeName(worksheet, employeeData['nombre'])  # 0 is the column index of the nombre
    _deleteRowAtSheet(worksheet, jobIdToFind_RowIndex)
    workbook.save(os.path.join(employeesFilePath, employeesFileName))

# Auxiliar functions

def readRowValues(worksheet, nRow, min_col, max_col):
    columnNames = []
    for row in worksheet.iter_cols(min_row=nRow, max_row=nRow, min_col=min_col, max_col=max_col):
        for cell in row:
            columnNames.append(cell.value)
    return columnNames



def _getRowIndexByEmployeeName(worksheet, employeeNameToFind):
    '''Looks for the index of the row whoose column named nombre contains employeeNameToFind

    :param worksheet: The excel worksheet readed
    :type worksheet: Worksheet
    :param employeeNameToFind: The name to find in the column named nombre
    :type employeeNameToFind: string

    :returns:   The index of the row which contains the employeeNameToFind in the column named nombre.
                If there is not row which contains the employeeNameToFind the returns -1
    :rtype: int
    '''

    employeeNameToFind_RowIndex = -1

    for rowIndex in range(1, worksheet.max_row):  # row_count in excel starts with 1
        columnValues = readRowValues(worksheet, rowIndex + 1, 1,None)  # read the other rows. Row index + 1 because we skip the header
        if employeeNameToFind == columnValues[0]:
            employeeNameToFind_RowIndex = rowIndex + 1
            break

    return employeeNameToFind_RowIndex


def _insertRowAtSheet(worksheet, rowToInsert, employeeData):
    columnNames = readRowValues(worksheet, 1, 1, None)
    worksheet.insert_rows(idx=rowToInsert)
    for columnName_Index, columnName in enumerate(columnNames):
        worksheet.cell(row=rowToInsert, column=columnName_Index+1).value = employeeData[columnName]

def _deleteRowAtSheet(worksheet, rowToDelete_Index):
    worksheet.delete_rows(rowToDelete_Index, amount=1)

